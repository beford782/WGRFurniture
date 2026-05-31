#!/usr/bin/env python3
"""DreamFinder Store Data Converter - full bundle (S2-S6).

Reads a completed onboarding workbook (.xlsx) and emits the data files the live
app consumes:

    <output-dir>/data/mattresses.csv        (English - the live CSV contract)   [S2]
    <output-dir>/data/mattresses-es.csv     (Spanish - only when ES content)     [S2]
    <output-dir>/data/store-config.json     (17 committed top-level keys)         [S3/S6]
    <output-dir>/data/accessories.json      (bilingual array, preserved order)    [S3]
    <output-dir>/data/allowed-hosts.js      (M1 domain-lock allowlist)            [S6]
    <output-dir>/images/...                 (normalized JPG, with --source-images) [S4]
    <output-dir>/manifest.json              (PWA manifest)                        [S5]

and, unless skipped, shells out to <output-dir>/build-data.ps1 to regenerate
<output-dir>/data/mattresses.json from the CSVs (the existing, trusted path).

Usage (run from the repo root, or anywhere):

    python tools/convert_store_data.py <workbook.xlsx> [--output-dir DIR]
           [--build-json | --skip-build-json]

Design (docs/phase0-onboarding-pipeline-spec-2026-05-31.md sections 3/4):
  * Mattresses CSVs are header-driven (the Mattresses tab EN headers ARE the live
    CSV contract). store-config.json / accessories.json are built from the shared
    tools/workbook_schema.py column->path mapping (single source of truth, also
    used by the fixture generator and the future create_template rewrite).
  * store-config.allowedHosts is the M1 source of truth; it is projected into
    data/allowed-hosts.js (window.__DF_ALLOWED_HOSTS) which index.html loads
    synchronously before its domain-lock IIFE. rsaList defaults to [].
  * Image normalization (S4) requires --source-images; mattresses.json is built
    by build-data.ps1 (the trusted path), not written here.
  * build-data.ps1 runs from <output-dir> (its $PSScriptRoot scopes it to the
    output workspace), so it never touches the repo's data/ unless --output-dir is
    the repo itself.

Dependencies: stdlib + openpyxl. ASCII-only console output.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import shutil
import subprocess
import sys

# The shared schema + validation live alongside this file in tools/.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import workbook_schema as schema  # noqa: E402
import validation  # noqa: E402

import openpyxl  # noqa: E402

ES_SUFFIX = " (ES)"

# Store Info columns whose store-config value is a list (workbook holds them as a
# single comma-separated cell).
STORE_INFO_LISTS = {"languages", "allowedHosts"}

# Top-level store-config.json key order (committed Bel order) - readability only;
# canonical comparison is parse-based and order-insensitive.
STORE_CONFIG_KEY_ORDER = [
    "storeName", "storeKey", "languages", "logo", "colors", "gasUrl",
    "publicAssetRoot", "allowedHosts", "brands", "rsaList", "text", "text_es",
    "discount", "voice", "voice_es", "salesNotes", "salesNotes_es",
]

# Per-accessory key order (committed Bel order) - readability only.
ACCESSORY_KEY_ORDER = [
    "id", "name", "category", "price", "image", "description",
    "subType", "matchTags", "matchScores",
]

# manifest.json (S5). Key order matches committed; display/orientation are
# constants (kiosk app), not workbook columns. No icons key (committed has none).
MANIFEST_KEY_ORDER = [
    "name", "short_name", "description", "start_url",
    "display", "orientation", "background_color", "theme_color",
]
MANIFEST_CONSTANTS = {"display": "standalone", "orientation": "landscape"}

# Image normalization (S4). Source images are accepted in any of these formats and
# re-encoded to JPG. WebP output is intentionally NOT produced (Outlook desktop /
# iOS Mail render WebP unreliably in result emails - CLAUDE.md image convention).
SOURCE_IMAGE_EXTS = (".jpg", ".jpeg", ".png", ".webp")
IMAGE_LONG_EDGE = 1000
# TODO(deps): image normalization requires Pillow, which is not yet declared in a
# requirements.txt / pyproject. Declare it (alongside openpyxl) in a later pass.


def _s(value) -> str:
    """Cell -> string. Blank for None; never coerce numbers when stringifying."""
    if value is None:
        return ""
    return value if isinstance(value, str) else str(value)


def _blank(value) -> bool:
    return value is None or str(value).strip() == ""


def set_path(d: dict, dotted: str, value) -> None:
    """Set a nested value by dotted path, creating intermediate dicts."""
    parts = dotted.split(".")
    cur = d
    for p in parts[:-1]:
        cur = cur.setdefault(p, {})
    cur[parts[-1]] = value


def read_tab(wb, name):
    """Return (headers, rows). Values are raw (types preserved). Rows that are
    entirely blank are skipped. Raises if the tab is missing."""
    if name not in wb.sheetnames:
        raise SystemExit(f"ERROR: workbook has no {name!r} tab "
                         f"(found: {', '.join(wb.sheetnames)})")
    ws = wb[name]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [h for h in header_row if h is not None]
    rows = []
    for vals in ws.iter_rows(min_row=2, values_only=True):
        d = {h: (vals[i] if i < len(vals) else None) for i, h in enumerate(headers)}
        if any(not _blank(v) for v in d.values()):
            rows.append(d)
    return headers, rows


# -- Mattresses (S2) ----------------------------------------------------------

def write_csv(path, fieldnames, rows):
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in fieldnames})


def emit_mattress_csvs(headers, rows, data_dir):
    """Write data/mattresses.csv (EN) and, when ES content exists,
    data/mattresses-es.csv. Returns (en_path, es_path_or_None)."""
    en_headers = [h for h in headers if not h.endswith(ES_SUFFIX)]
    es_headers = [h for h in headers if h.endswith(ES_SUFFIX)]
    mrows = [r for r in rows if _s(r.get("id")).strip()]

    en_path = os.path.join(data_dir, "mattresses.csv")
    write_csv(en_path, en_headers, [{h: _s(r.get(h)) for h in en_headers} for r in mrows])

    es_path = None
    if es_headers:
        es_names = [h[:-len(ES_SUFFIX)] for h in es_headers]
        has_es = any(_s(r.get(h)).strip() for r in mrows for h in es_headers)
        if has_es:
            es_fieldnames = ["id"] + es_names
            es_rows = []
            for r in mrows:
                er = {"id": _s(r.get("id"))}
                for h, name in zip(es_headers, es_names):
                    er[name] = _s(r.get(h))
                es_rows.append(er)
            es_path = os.path.join(data_dir, "mattresses-es.csv")
            write_csv(es_path, es_fieldnames, es_rows)
    return en_path, es_path


# -- store-config.json (S3) ---------------------------------------------------

def build_sales_notes(wb):
    """Build (salesNotes, salesNotes_es) from the SalesNotes tab."""
    _, rows = read_tab(wb, "SalesNotes")
    sn = {"subBrands": {}, "brands": {}}
    sn_es = {"subBrands": {}, "brands": {}}
    for r in rows:
        typ = _s(r.get("Type")).strip()
        key = _s(r.get("Key")).strip()
        if not typ or not key:
            continue
        if typ == "subBrand":
            fmt = _s(r.get("Format")).strip()
            if fmt == "full":
                sn["subBrands"][key] = {
                    "format": "full",
                    "lead": _s(r.get("Lead")),
                    "demo": _s(r.get("Demo")),
                    "close": _s(r.get("Close")),
                }
                es = {}
                for jk, hk in (("lead", "Lead (ES)"), ("demo", "Demo (ES)"),
                               ("close", "Close (ES)")):
                    if not _blank(r.get(hk)):
                        es[jk] = _s(r.get(hk))
                if es:
                    sn_es["subBrands"][key] = es
            elif fmt == "coaching":
                sn["subBrands"][key] = {"format": "coaching", "rsaNote": _s(r.get("RSA Note"))}
                if not _blank(r.get("RSA Note (ES)")):
                    sn_es["subBrands"][key] = {"rsaNote": _s(r.get("RSA Note (ES)"))}
        elif typ == "brand":
            sn["brands"][key] = {"story": _s(r.get("Story"))}
            if not _blank(r.get("Story (ES)")):
                sn_es["brands"][key] = {"story": _s(r.get("Story (ES)"))}
    return sn, sn_es


def build_store_config(wb):
    """Assemble store-config.json from Store Info + Brands + SalesNotes tabs."""
    _, si_rows = read_tab(wb, "Store Info")
    si = si_rows[0] if si_rows else {}

    cfg = {}
    for col in schema.get_columns("Store Info"):
        key = col.key
        if key.startswith("manifest."):
            continue  # manifest.json is S5
        cell = si.get(col.name)
        if key in STORE_INFO_LISTS:
            value = [s.strip() for s in _s(cell).split(",") if s.strip()]
        else:
            value = "" if cell is None else cell  # preserve numeric types
        set_path(cfg, key, value)

    # Brands tab -> brands[]. Workbook holds the logo file name; config stores the
    # images/brands/<file> path.
    _, brand_rows = read_tab(wb, "Brands")
    brands = []
    for r in brand_rows:
        name = _s(r.get("Brand Name")).strip()
        if not name:
            continue
        logo_file = _s(r.get("Logo File Name")).strip()
        brands.append({
            "name": name,
            "logo": f"images/brands/{logo_file}" if logo_file else "",
        })
    cfg["brands"] = brands

    cfg["rsaList"] = []  # runtime-populated; default empty

    sn, sn_es = build_sales_notes(wb)
    cfg["salesNotes"] = sn
    cfg["salesNotes_es"] = sn_es

    # Reorder top-level keys to the committed order (readability only).
    return {k: cfg[k] for k in STORE_CONFIG_KEY_ORDER if k in cfg}


# -- manifest.json (S5) -------------------------------------------------------

def build_manifest(wb):
    """Assemble manifest.json from the Store Info manifest.* columns + the
    display/orientation constants. (manifest.* are skipped by build_store_config,
    so the two outputs do not overlap.) No icons key (committed has none)."""
    _, si_rows = read_tab(wb, "Store Info")
    si = si_rows[0] if si_rows else {}
    values = dict(MANIFEST_CONSTANTS)
    for col in schema.get_columns("Store Info"):
        if col.key.startswith("manifest."):
            sub = col.key[len("manifest."):]
            cell = si.get(col.name)
            values[sub] = "" if cell is None else cell  # preserve string/type
    return {k: values[k] for k in MANIFEST_KEY_ORDER if k in values}


# -- accessories.json (S3) ----------------------------------------------------

def build_accessories(wb):
    """Build the accessories array (preserved tab/row order)."""
    _, rows = read_tab(wb, "Accessories")
    out = []
    for r in rows:
        if _blank(r.get("ID")):
            continue
        acc = {}
        for col in schema.get_columns("Accessories"):
            key = col.key
            cell = r.get(col.name)
            if key == "matchTags":
                acc["matchTags"] = [t.strip() for t in _s(cell).split(",") if t.strip()]
            elif key.startswith("matchScores."):
                if not _blank(cell):
                    set_path(acc, key, cell)  # numeric preserved
            elif key == "subType":
                if not _blank(cell):
                    acc["subType"] = cell
            elif key == "price":
                acc["price"] = cell  # numeric preserved
            elif key == "id":
                acc["id"] = _s(cell)
            else:
                # name.en / category.es / description.en / image -> verbatim
                set_path(acc, key, "" if cell is None else cell)
        acc.setdefault("matchTags", [])
        acc.setdefault("matchScores", {})
        out.append({k: acc[k] for k in ACCESSORY_KEY_ORDER if k in acc})
    return out


def write_json(path, obj):
    with open(path, "w", encoding="utf-8", newline="") as f:
        json.dump(obj, f, indent=2, ensure_ascii=False)
        f.write("\n")


def write_allowed_hosts_js(path, hosts):
    """Project store-config.allowedHosts into the M1 domain-lock allowlist JS
    (loaded synchronously by index.html before the main script). JSON serialization
    keeps the array JS-safe; trailing semicolon + newline."""
    with open(path, "w", encoding="utf-8", newline="") as f:
        f.write("window.__DF_ALLOWED_HOSTS = " + json.dumps(hosts) + ";\n")


# -- image normalization (S4) -------------------------------------------------

def _index_source_images(src_dir):
    """Map lowercased filename stem -> source path, for supported image types."""
    idx = {}
    for fn in os.listdir(src_dir):
        stem, ext = os.path.splitext(fn)
        if ext.lower() in SOURCE_IMAGE_EXTS:
            idx[stem.lower()] = os.path.join(src_dir, fn)
    return idx


def _normalize_one(Image, src_path, dst_path, quality):
    """Re-encode one source image to JPG (RGB, long-edge<=cap, given quality)."""
    img = Image.open(src_path)
    if img.mode in ("RGBA", "LA"):
        bg = Image.new("RGB", img.size, (255, 255, 255))
        bg.paste(img, mask=img.split()[-1])
        img = bg
    elif img.mode != "RGB":
        img = img.convert("RGB")
    w, h = img.size
    if max(w, h) > IMAGE_LONG_EDGE:
        if w >= h:
            img = img.resize((IMAGE_LONG_EDGE, int(h * IMAGE_LONG_EDGE / w)), Image.LANCZOS)
        else:
            img = img.resize((int(w * IMAGE_LONG_EDGE / h), IMAGE_LONG_EDGE), Image.LANCZOS)
    img.save(dst_path, "JPEG", quality=quality, optimize=True)


def normalize_images(source_root, output_dir, mattress_stems, accessory_stems, quality):
    """Normalize product images to JPG into <output-dir>/images/{mattresses,
    accessories}/. Requires Pillow (lazy import). Output stems:
      * mattresses  -> lower(name)         (matches build-data.ps1's resolution)
      * accessories -> basename(image cell) (matches accessories.json; NOT id)
    Source images are matched case-insensitively by stem. A missing source image
    is a hard error (bad onboarding asset). Brand/store/PWA images are NOT touched."""
    try:
        from PIL import Image
    except ImportError:
        raise SystemExit(
            "ERROR: image normalization (--source-images) requires Pillow.\n"
            "       Install it with: pip install Pillow\n"
            "       (Or omit --source-images / pass --skip-image-normalization to "
            "emit CSV/JSON only.)")

    total = 0
    for subdir, stems in (("mattresses", mattress_stems),
                          ("accessories", accessory_stems)):
        if not stems:
            continue
        src_dir = os.path.join(source_root, subdir)
        if not os.path.isdir(src_dir):
            raise SystemExit(f"ERROR: source image folder not found: {src_dir}")
        idx = _index_source_images(src_dir)
        out_dir = os.path.join(output_dir, "images", subdir)
        os.makedirs(out_dir, exist_ok=True)
        for stem in stems:
            src = idx.get(stem.lower())
            if not src:
                raise SystemExit(
                    f"ERROR: no source image for {subdir}/{stem} "
                    f"(looked for {stem}.[jpg|jpeg|png|webp] in {src_dir})")
            _normalize_one(Image, src, os.path.join(out_dir, stem + ".jpg"), quality)
            total += 1
        print(f"  normalized {len(stems)} {subdir} image(s) -> {out_dir}")
    return total


# -- build-data.ps1 (mattresses.json) ----------------------------------------

def run_build_data(output_dir):
    """Invoke <output-dir>/build-data.ps1 to regenerate mattresses.json.
    Warn + skip (never fail the emit) if PowerShell or the script is absent.
    Returns True iff build-data.ps1 ran and produced data/mattresses.json."""
    script = os.path.join(output_dir, "build-data.ps1")
    if not os.path.exists(script):
        print(f"[build-json] skipped: {script} not found.")
        return False
    ps = shutil.which("pwsh") or shutil.which("powershell")
    if not ps:
        print("[build-json] skipped: no pwsh/powershell on PATH.")
        return False
    print(f"[build-json] running {os.path.basename(ps)} -File {script}")
    proc = subprocess.run(
        [ps, "-NoProfile", "-ExecutionPolicy", "Bypass", "-File", script],
        capture_output=True, text=True)
    if proc.stdout.strip():
        print("  " + proc.stdout.strip().replace("\n", "\n  "))
    if proc.returncode != 0:
        print(f"[build-json] WARNING: build-data.ps1 exited {proc.returncode} "
              f"(CSV/JSON output is still valid).")
        if proc.stderr.strip():
            print("  " + proc.stderr.strip().replace("\n", "\n  "))
        return False
    return os.path.exists(os.path.join(output_dir, "data", "mattresses.json"))


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(
        description="Convert an onboarding workbook into DreamFinder data files.",
        formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("workbook", help="Path to the onboarding .xlsx")
    parser.add_argument("--output-dir", default=".",
                        help="Where to write data/ (default: current directory)")
    parser.add_argument("--source-images", default=None,
                        help="Folder with mattresses/ and accessories/ subdirs of raw "
                             "images; when set, normalize them to JPG into "
                             "<output-dir>/images/. Omit to skip image normalization.")
    parser.add_argument("--skip-image-normalization", action="store_true",
                        help="Do not normalize images even if --source-images is given.")
    parser.add_argument("--image-quality", type=int, default=88,
                        help="JPEG quality for normalized images (default 88).")
    parser.add_argument("--no-validate", action="store_true",
                        help="Skip input validation (not recommended).")
    parser.add_argument("--validate-only", action="store_true",
                        help="Validate the workbook and exit; write no files.")
    parser.add_argument("--warnings-as-errors", action="store_true",
                        help="Treat validation warnings as blocking errors.")
    parser.add_argument("--require-gas-url", action="store_true",
                        help="Treat a blank/placeholder gasUrl as a blocking error.")
    group = parser.add_mutually_exclusive_group()
    group.add_argument("--build-json", action="store_true",
                       help="Run build-data.ps1 to regenerate mattresses.json (default).")
    group.add_argument("--skip-build-json", action="store_true",
                       help="Do not invoke build-data.ps1.")
    args = parser.parse_args(argv)

    do_validate = not args.no_validate
    data_dir = os.path.join(args.output_dir, "data")

    print(f"Reading {args.workbook}...")
    wb = openpyxl.load_workbook(args.workbook, read_only=True, data_only=True)
    report = validation.ValidationReport()
    try:
        # Read present tabs (guarded so a missing tab is a validation error, not a
        # crash). raw_tabs maps present tab name -> (headers, rows).
        raw_tabs = {name: read_tab(wb, name) for name in schema.get_tab_names()
                    if name in wb.sheetnames}

        # Structure gate: if the workbook is structurally unsound we cannot safely
        # assemble the bundle, so abort here before writing anything.
        if do_validate:
            report.merge(validation.validate_structure(raw_tabs))
            if not report.ok:
                print(report.summary())
                print("[validate] blocking structure errors - no files written.")
                return 1

        # Assemble everything in memory (no writes yet).
        m_headers, m_rows = raw_tabs.get("Mattresses") or read_tab(wb, "Mattresses")
        config = build_store_config(wb)
        accessories = build_accessories(wb)
        manifest = build_manifest(wb)
    finally:
        wb.close()

    # Value + catalog validation (store-config, mattresses, accessories,
    # SalesNotes, and source-image existence when --source-images is provided).
    if do_validate:
        langs = config.get("languages")
        report.merge(validation.validate_store_config(
            config, manifest, require_gas_url=args.require_gas_url))
        report.merge(validation.validate_mattresses(
            raw_tabs, source_images=args.source_images,
            skip_images=args.skip_image_normalization, languages=langs))
        report.merge(validation.validate_accessories(
            raw_tabs, source_images=args.source_images,
            skip_images=args.skip_image_normalization, languages=langs))
        report.merge(validation.validate_sales_notes(raw_tabs, languages=langs))
        print(report.summary())
        blocking = report.blocking(warnings_as_errors=args.warnings_as_errors)
        if args.validate_only:
            return 0 if not blocking else 1
        if blocking:
            print("[validate] blocking errors - no files written.")
            return 1
    elif args.validate_only:
        print("[validate] --validate-only with --no-validate: nothing to check.")
        return 0

    # Write phase (validation passed or skipped).
    os.makedirs(data_dir, exist_ok=True)
    en_path, es_path = emit_mattress_csvs(m_headers, m_rows, data_dir)
    print(f"  wrote {en_path} ({sum(1 for r in m_rows if _s(r.get('id')).strip())} rows)")
    if es_path:
        print(f"  wrote {es_path}")
    else:
        print("  no Spanish content - data/mattresses-es.csv omitted")

    cfg_path = os.path.join(data_dir, "store-config.json")
    write_json(cfg_path, config)
    print(f"  wrote {cfg_path} ({len(config)} top-level keys)")

    ah_path = os.path.join(data_dir, "allowed-hosts.js")
    write_allowed_hosts_js(ah_path, config.get("allowedHosts", []))
    print(f"  wrote {ah_path} ({config.get('allowedHosts', [])})")

    acc_path = os.path.join(data_dir, "accessories.json")
    write_json(acc_path, accessories)
    print(f"  wrote {acc_path} ({len(accessories)} items)")

    man_path = os.path.join(args.output_dir, "manifest.json")  # repo root, not data/
    write_json(man_path, manifest)
    print(f"  wrote {man_path} ({len(manifest)} keys)")

    # Image normalization (S4) - optional; product images only.
    if args.source_images and not args.skip_image_normalization:
        mattress_stems = [
            _s(r.get("name")).strip().lower()
            for r in m_rows
            if _s(r.get("id")).strip() and _s(r.get("name")).strip()
        ]
        accessory_stems = [
            os.path.splitext(os.path.basename(a["image"]))[0]
            for a in accessories if a.get("image")
        ]
        print(f"Normalizing images from {args.source_images} "
              f"(quality {args.image_quality})...")
        normalize_images(args.source_images, args.output_dir,
                         mattress_stems, accessory_stems, args.image_quality)
    elif args.source_images and args.skip_image_normalization:
        print("[images] skipped (--skip-image-normalization).")

    built = False
    if not args.skip_build_json:
        built = run_build_data(args.output_dir)
    else:
        print("[build-json] skipped (--skip-build-json).")

    # Post-emit validation: verify the bundle we just wrote (V3). Generated files
    # are NOT deleted on failure (useful for debugging a post-emit problem).
    if do_validate:
        post = validation.validate_generated_outputs(
            args.output_dir, build_json=built, languages=config.get("languages"))
        if post.errors or post.warnings:
            print(post.summary())
        if post.blocking(warnings_as_errors=args.warnings_as_errors):
            print("[validate] post-emit validation failed (files were written; "
                  "see above).")
            return 1

    print("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
